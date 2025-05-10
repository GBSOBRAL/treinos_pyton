import openpyxl
import customtkinter as ctk

book = openpyxl.load_workbook("Dados.xlsx")
page_Emp = book["Empresas"]
page_Plays = book["Plays"]
page_TP = book["Touchpoints"]
page_Ativ = book["Atividades"]
page_Resp = book["Responsável"]
page_Cont = book["Contato"]
page_Import = book["Importação"]

line = []

try:
    for row_Ativ in page_Ativ.iter_rows(min_row = 2):
        line = [row_Ativ[0].value, row_Ativ[1].value]
        for row_Emp in page_Emp.iter_rows(min_row = 2):
                if str(row_Emp[1].value) in line[1]:
                        line = line + [row_Emp[0].value, row_Emp[2].value]
                        for row_Cont in page_Cont.iter_rows(min_row = 2):
                                if str(row_Cont[1].value) == line[2]:
                                        line = line + [row_Cont[0].value]
                                        page_Import.append(line)
                                        print(line)

except:
       pass

book.save("Dados.xlsx")