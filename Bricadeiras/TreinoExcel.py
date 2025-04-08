import sys
import math
import random
import threading
import time
import openpyxl
from functools import reduce

#Criar um book(arquivo de planilhas)
book = openpyxl.Workbook()

#Visualizar páginas existentes
print(book.sheetnames)

#Criar uma página
book.create_sheet("Frutas")

#Selecionar página para edição
frutas_page = book["Frutas"]

#inserir dados
frutas_page.append(["Nome", "Quantidade", "Preço"])
frutas_page.append(["Bananas", "5", "R$1.99"])
frutas_page.append(["Maçã", "2", "R$2.99"])
frutas_page.append(["Uva", "10", "R$4.49"])
frutas_page.append(["Mamão", "7", "R$5.99"])

#Salvar planilha
book.save("Frutaria.xlsx")

