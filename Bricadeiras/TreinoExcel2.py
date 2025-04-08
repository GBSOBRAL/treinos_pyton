import sys
import math
import random
import threading
import time
from traceback import print_tb

import openpyxl
from functools import reduce

book = openpyxl.load_workbook("Frutaria.xlsx")

frutas_page = book["Frutas"]

for rows in frutas_page.iter_rows(min_row = 1, max_row=5):
    for cell in rows:
        if cell.value == "Bananas":
            cell.value = "Cereja"

for rows in frutas_page.iter_rows(min_row = 1, max_row=5):
        print(rows[0].value, rows[1].value, rows[2].value)

book.save("Frutaria.xlsx")