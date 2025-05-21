import openpyxl

book = openpyxl.Workbook()

#Criar uma página
book.create_sheet("Computadores")

#Selecionar página para edição
comp_page = book["Computadores"]

#inserir dados
comp_page.append(["Eletrônica", "Memória RAM", "Preço"])
comp_page.append(["Computador 1", "8GB RAM", "R$2.500"])
comp_page.append(["Computador 2", "16GB RAM", "R$5.500"])
comp_page.append(["Computador 3", "32GB RAM", "R$10.550"])

#Salvar planilha
book.save("LojaComputadores.xlsx")

#Mostrar páginas
print(book.sheetnames)

