from openpyxl import load_workbook, Workbook

#Selecionar Arquivo e Pagina
plan_vendas = load_workbook("vendas_de_lanches.xlsx")
pag_vendas = plan_vendas["Sheet"]

#Ler Planilha
for linha in pag_vendas.iter_rows(values_only=True):
    print(linha)

#Automatizar entrada de dados / Inserir dados
#Criando arquivo
plan_contas = Workbook()
pag1 = plan_contas.active

#Lendo e populando planilha com arquivo de texto
with open("anotações.txt","r",encoding="utf-8") as arquivo:
    for linha in arquivo:
        pag1.append(linha.split(","))
plan_contas.save("contas_a_pagar.xlsx")

#Enviar os dados por e-mail/wpp
